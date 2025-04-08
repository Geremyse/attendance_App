from flask import Flask, render_template, redirect, url_for, flash, request, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField, SelectField, FileField
from wtforms.validators import DataRequired, Length, EqualTo
import mysql.connector
import cv2
import face_recognition
from datetime import datetime
from flask_socketio import SocketIO, emit
import base64
import threading
from functools import wraps
from flask import abort
from forms import LoginForm, RegisterForm, AddStudentForm, TeacherRegistrationForm, StudentRegistrationForm, TeacherEditForm, ParentRegistrationForm
import os
import json
import random
import shutil
from werkzeug.utils import secure_filename
import pandas as pd
from collections import defaultdict
import re
from flask_bcrypt import Bcrypt
from flask import send_file

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
socketio = SocketIO(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

UPLOAD_FOLDER = 'known_faces'
UPLOAD_FOLDER_FOR_PROFILES_S = 'static/students_photo'
UPLOAD_FOLDER_FOR_PROFILES_T = 'static/teachers_photo'

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['UPLOAD_FOLDER_FOR_PROFILES_S'] = UPLOAD_FOLDER_FOR_PROFILES_S  # Папка для сохранения фотографий профиля учеников
app.config['UPLOAD_FOLDER_FOR_PROFILES_T'] = UPLOAD_FOLDER_FOR_PROFILES_T  # Папка для сохранения фотографий профиля учителей

bcrypt = Bcrypt(app)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS, {'xlsx'}


def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="attendance_db"
    )

def load_known_faces():
    print("hello world")
    known_faces_dir = 'known_faces'
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT CONCAT(last_name, ' ', first_name, ' ', middle_name), face_encoding FROM students")
    students = cursor.fetchall()
    cursor.close()
    conn.close()

    known_face_encodings = []
    known_face_names = []

    for name, face_encoding in students:
        filename = f"{face_encoding}.jpg"
        file_path = os.path.join(known_faces_dir, filename)
        if os.path.exists(file_path):
            image = face_recognition.load_image_file(file_path)
            encodings = face_recognition.face_encodings(image)
            if encodings:
                known_face_encodings.append(encodings[0])
                known_face_names.append(name)
            else:
                print(f"No face found in {filename}")
        else:
            print(f"File {filename} not found")

    return known_face_encodings, known_face_names

load_known_faces()

# Глобальная переменная для управления захватом видео
capture_active = False

# Захват изображений и распознавание лиц

import cv2
import face_recognition
import base64
from PIL import Image, ImageDraw, ImageFont
import numpy as np

def capture_and_recognize(teacher_id):
    global capture_active
    video_capture = cv2.VideoCapture(1)
    video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 640)  # Уменьшение разрешения
    video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)  # Уменьшение разрешения

    frame_count = 0
    process_every_n_frames = 5  # Обработка каждого пятого кадра

    known_face_encodings, known_face_names = load_known_faces()

    while capture_active:
        ret, frame = video_capture.read()
        if not ret:
            print("Failed to capture frame")
            break

        frame_count += 1
        if frame_count % process_every_n_frames == 0:
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            face_locations = face_recognition.face_locations(rgb_frame)
            face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Unknown"

                if True in matches:
                    first_match_index = matches.index(True)
                    name = known_face_names[first_match_index]
                    print(f"Recognized: {name}")
                    record_attendance(name, teacher_id)
                else:
                    print(f"Unknown face detected")

                # Отображение имени на изображении
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)

                # Использование Pillow для рендеринга текста с поддержкой кириллицы
                pil_image = Image.fromarray(frame)
                draw = ImageDraw.Draw(pil_image)
                font = ImageFont.truetype("arial.ttf", 20)  # Убедитесь, что у вас есть шрифт, поддерживающий кириллицу
                draw.text((left + 6, bottom - 30), name, font=font, fill=(255, 255, 255))
                frame = np.array(pil_image)

            # Преобразование кадра в base64 для передачи через WebSocket
            _, buffer = cv2.imencode('.jpg', frame)
            frame_base64 = base64.b64encode(buffer).decode('utf-8')
            socketio.emit('frame', {'image': frame_base64})

    video_capture.release()
    cv2.destroyAllWindows()


# Запись данных о посещаемости
def get_day_of_week_in_russian(day_of_week_english):
    days_mapping = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье'
    }
    return days_mapping.get(day_of_week_english, day_of_week_english)

def record_attendance(student_name, teacher_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM students WHERE CONCAT(last_name, ' ', first_name, ' ', middle_name) = %s", (student_name,))
    student = cursor.fetchone()
    if student is None:
        print(f"Student with name {student_name} not found in the database.")
        cursor.close()
        conn.close()
        return

    student_id = student[0]
    current_time = datetime.now()
    current_date = current_time.strftime('%Y-%m-%d')
    current_time_str = current_time.strftime('%H:%M:%S')
    current_day_of_week_english = current_time.strftime('%A')  # Получаем текущий день недели на английском
    current_day_of_week = get_day_of_week_in_russian(current_day_of_week_english)  # Переводим на русский

    # Определение номера урока на основе времени и дня недели
    cursor.execute("SELECT lesson_number, time FROM schedule WHERE day_of_week = %s", (current_day_of_week,))
    schedule_rows = cursor.fetchall()
    lesson_number = None

    for row in schedule_rows:
        time_range = row[1].strip()  # Удаляем лишние пробелы
        start_time_str, end_time_str = time_range.split('-')
        start_time_str = start_time_str.strip()  # Удаляем лишние пробелы
        end_time_str = end_time_str.strip()  # Удаляем лишние пробелы

        try:
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            end_time = datetime.strptime(end_time_str, '%H:%M').time()
            current_time_obj = datetime.strptime(current_time_str, '%H:%M:%S').time()

            if start_time <= current_time_obj <= end_time:
                lesson_number = row[0]
                break
        except ValueError as e:
            print(f"Error parsing time range {time_range}: {e}")

    if lesson_number is None:
        print(f"No lesson found for time {current_time_str} on day {current_day_of_week}")
        cursor.close()
        conn.close()
        return

    # Проверка, была ли уже записана посещаемость для данного ученика в текущий день и урок
    cursor.execute("SELECT * FROM attendance WHERE student_id = %s AND DATE(attendance_time) = %s AND lesson_number = %s", (student_id, current_date, lesson_number))
    existing_record = cursor.fetchone()

    if not existing_record:
        query = "INSERT INTO attendance (student_id, attendance_time, responsible_teacher_id, lesson_number) VALUES (%s, %s, %s, %s)"
        values = (student_id, current_time, teacher_id, lesson_number)
        cursor.execute(query, values)
        conn.commit()
        print(f"Attendance recorded for student {student_name} at {current_time}")

        # Преобразование current_time в строку перед отправкой через WebSocket
        current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
        socketio.emit('attendance_update', {'data': [(student_name, current_time_str)]})

    cursor.close()
    conn.close()



@socketio.on('connect')
def handle_connect():
    print('Client connected')

@socketio.on('disconnect')
def handle_disconnect():
    print('Client disconnected')

@socketio.on('get_attendance_updates')
def handle_get_attendance_updates():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name), attendance.attendance_time FROM attendance JOIN students ON attendance.student_id = students.id")
    attendance_data = cursor.fetchall()
    cursor.close()
    conn.close()

    # Преобразование данных в список словарей для удобства
    attendance_list = []
    for name, attendance_time in attendance_data:
        attendance_list.append({
            'name': name,
            'attendance_time': attendance_time.strftime('%Y-%m-%d %H:%M:%S')  # Преобразование datetime в строку
        })

    socketio.emit('attendance_update', {'data': attendance_list})


# Модель пользователя
""" class User(UserMixin):
    def __init__(self, id, username, password, role, last_name=None, first_name=None, middle_name=None):
        self.id = id
        self.username = username
        self.password = password
        self.role = role
        self.last_name = last_name
        self.first_name = first_name
        self.middle_name = middle_name

    def is_admin(self):
        return self.role == 'admin'

    def is_teacher(self):
        return self.role == 'teacher'

    def is_parent(self):
        return self.role == 'parent' """

class User(UserMixin):
    def __init__(self, id, username, password, role, teacher_id=None, parent_id=None, last_name=None, first_name=None, middle_name=None):
        self.id = id
        self.username = username
        self.password = password
        self.role = role
        self.teacher_id = teacher_id
        self.parent_id = parent_id
        self.last_name = last_name
        self.first_name = first_name
        self.middle_name = middle_name

    def is_admin(self):
        return self.role == 'admin'

    def is_teacher(self):
        return self.role == 'teacher'

    def is_parent(self):
        return self.role == 'parent'


""" @login_manager.user_loader
def load_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, password, role, teacher_id, parent_id FROM users WHERE id = %s", (user_id,))
    user = cursor.fetchone()
    if user:
        role = user[3]
        if role == 'teacher':
            cursor.execute("SELECT last_name, first_name, middle_name FROM teachers WHERE id = %s", (user[4],))
            teacher_data = cursor.fetchone()
            if teacher_data:
                last_name, first_name, middle_name = teacher_data
            else:
                last_name, first_name, middle_name = None, None, None
        elif role == 'parent':
            # Предположим, что у родителей есть таблица parents с аналогичными полями
            cursor.execute("SELECT last_name, first_name, middle_name FROM parents WHERE id = %s", (user[5],))
            parent_data = cursor.fetchone()
            if parent_data:
                last_name, first_name, middle_name = parent_data
            else:
                last_name, first_name, middle_name = None, None, None
        else:
            last_name, first_name, middle_name = None, None, None

        cursor.close()
        conn.close()
        return User(user[0], user[1], user[2], user[3], last_name, first_name, middle_name)
    return None """

@login_manager.user_loader
def load_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, password, role, teacher_id, parent_id FROM users WHERE id = %s", (user_id,))
    user = cursor.fetchone()
    if user:
        role = user[3]
        teacher_id = user[4]
        parent_id = user[5]
        last_name = first_name = middle_name = None

        if role == 'teacher':
            cursor.execute("SELECT last_name, first_name, middle_name FROM teachers WHERE id = %s", (teacher_id,))
            teacher_data = cursor.fetchone()
            if teacher_data:
                last_name, first_name, middle_name = teacher_data
        elif role == 'parent':
            cursor.execute("SELECT last_name, first_name, middle_name FROM parents WHERE id = %s", (parent_id,))
            parent_data = cursor.fetchone()
            if parent_data:
                last_name, first_name, middle_name = parent_data

        cursor.close()
        conn.close()
        return User(user[0], user[1], user[2], user[3], teacher_id, parent_id, last_name, first_name, middle_name)
    return None


def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return abort(403)
            if role == 'admin' and not current_user.is_admin():
                return abort(403)
            if role == 'teacher' and not current_user.is_teacher():
                return abort(403)
            if role == 'parent' and not current_user.is_parent():
                return abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/')
@login_required
def index():
    if current_user.is_admin():
        return redirect(url_for('admin_dashboard'))
    elif current_user.is_teacher():
        return redirect(url_for('teacher_dashboard'))
    elif current_user.is_parent():
        return redirect(url_for('parent_dashboard'))
    return redirect(url_for('login'))

@app.route('/admin')
@login_required
@role_required('admin')
def admin_dashboard():
    return render_template('admin_dashboard.html')


@app.route('/admin/teachers', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_teachers():
    form = TeacherRegistrationForm()
    form.class_id.choices = [(row[0], row[1]) for row in get_classes() if row[1] != 'No Class']
    if form.validate_on_submit():
        last_name = form.last_name.data
        first_name = form.first_name.data
        middle_name = form.middle_name.data
        username = form.username.data
        password = form.password.data
        has_classes = request.form.get('has_classes') == 'on'
        photo = form.photo.data

        # Хеширование пароля
        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Проверка уникальности username
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        existing_user = cursor.fetchone()
        if existing_user:
            flash('Имя пользователя уже существует', 'danger')
            return redirect(url_for('admin_teachers'))

        # Handle photo upload
        face_encoding = None
        if photo and allowed_file(photo.filename):
            face_encoding = random.randint(100000, 999999)  # Генерация случайного числа для face_encoding
            filename = f"{face_encoding}.jpg"
            photo_path = os.path.join(app.config['UPLOAD_FOLDER_FOR_PROFILES_T'], filename)
            photo.save(photo_path)

        cursor.execute("INSERT INTO teachers (last_name, first_name, middle_name, face_encoding) VALUES (%s, %s, %s, %s)",
                       (last_name, first_name, middle_name, face_encoding))
        teacher_id = cursor.lastrowid
        cursor.execute("INSERT INTO users (username, password, role, teacher_id) VALUES (%s, %s, %s, %s)",
                       (username, hashed_password, 'teacher', teacher_id))

        if has_classes:
            class_ids = form.class_id.data
            for class_id in class_ids:
                cursor.execute("INSERT INTO teacher_classes (teacher_id, class_id, has_classes) VALUES (%s, %s, %s)",
                               (teacher_id, class_id, 1))
        else:
            cursor.execute("INSERT INTO teacher_classes (teacher_id, class_id, has_classes) VALUES (%s, %s, %s)",
                           (teacher_id, 0, 0))  # Используем 0 для обозначения отсутствия класса

        conn.commit()
        cursor.close()
        conn.close()
        flash('Учитель добавлен успешно', 'success')
        return redirect(url_for('admin_teachers'))

    teachers = get_teachers_with_combined_classes()
    return render_template('admin_teachers.html', form=form, teachers=teachers)


@app.route('/admin/students', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_students():
    form = StudentRegistrationForm()
    form.class_id.choices = [(row[0], row[1]) for row in get_classes()]

    if form.validate_on_submit():
        last_name = form.last_name.data
        first_name = form.first_name.data
        middle_name = form.middle_name.data
        class_id = form.class_id.data
        photo = form.photo.data
        if photo and allowed_file(photo.filename):
            face_encoding = random.randint(100000, 999999)  # Генерация случайного числа для face_encoding
            filename = f"{face_encoding}.jpg"
            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            photo.save(photo_path)
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO students (last_name, first_name, middle_name, face_encoding, class_id) VALUES (%s, %s, %s, %s, %s)",
                           (last_name, first_name, middle_name, face_encoding, class_id))
            conn.commit()
            cursor.close()
            conn.close()
            flash('Ученик добавлен успешно', 'success')
            return redirect(url_for('admin_students'))
        else:
            flash('Неверный формат файла', 'danger')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in {field}: {error}", 'danger')

    students = get_students()
    return render_template('admin_students.html', form=form, students=students)


@app.route('/admin/classes', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_classes():
    if request.method == 'POST':
        class_name = request.form.get('class_name')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO classes (name) VALUES (%s)", (class_name,))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Класс добавлен успешно', 'success')
        return redirect(url_for('admin_classes'))

    classes = get_classes()
    return render_template('admin_classes.html', classes=classes)

def get_students():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT students.*, classes.name FROM students JOIN classes ON students.class_id = classes.id")
    students = cursor.fetchall()
    cursor.close()
    conn.close()

    # Отладочная информация
    print("Students data fetched from the database:")
    for student in students:
        print(student)

    return students

def get_teachers():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT teachers.id, teachers.last_name, teachers.first_name, teachers.middle_name,
               users.username, users.password, users.role,
               GROUP_CONCAT(classes.name ORDER BY classes.name SEPARATOR ', ') as class_names,
               MAX(teacher_classes.has_classes) as has_classes,
               teachers.face_encoding,
               teacher_classes.class_id,
               teachers.email, teachers.phone, teachers.education, teachers.experience, teachers.face_encoding
        FROM teachers
        JOIN users ON teachers.id = users.teacher_id
        LEFT JOIN teacher_classes ON teachers.id = teacher_classes.teacher_id
        LEFT JOIN classes ON teacher_classes.class_id = classes.id
        GROUP BY teachers.id, users.username, users.password, users.role, teacher_classes.class_id
    """)
    teachers = cursor.fetchall()
    cursor.close()
    conn.close()
        # Отладочная информация
    print("Teachers data fetched from the database:")
    for teacher in teachers:
        print(teacher)

    return teachers


@app.route('/teacher')
@login_required
@role_required('teacher')
def teacher_dashboard():
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)  # Используем буферизованный курсор
    teacher_id = current_user.teacher_id

    # Получаем данные о посещаемости для учеников, за съемку которых отвечал текущий учитель
    cursor.execute("""
        SELECT CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name),
               classes.name, attendance.attendance_time, attendance.lesson_number
        FROM attendance
        JOIN students ON attendance.student_id = students.id
        JOIN classes ON students.class_id = classes.id
        WHERE attendance.responsible_teacher_id = %s
        ORDER BY attendance.attendance_time DESC
        LIMIT 5
    """, (teacher_id,))
    attendance_data = cursor.fetchall()

    # Проверяем, есть ли у учителя классы
    cursor.execute("""
        SELECT COUNT(*)
        FROM teacher_classes
        WHERE teacher_id = %s AND has_classes = 1
    """, (teacher_id,))
    has_classes = cursor.fetchone()[0] > 0

    # Получаем дополнительную информацию из расписания
    updated_attendance_data = []
    for record in attendance_data:
        lesson_number = record[3]
        attendance_time = record[2]
        current_day_of_week_english = attendance_time.strftime('%A')  # Получаем день недели по дате посещения на английском
        current_day_of_week = get_day_of_week_in_russian(current_day_of_week_english)  # Переводим на русский

        # Отладочный вывод в консоль
        print(f"Query: SELECT day_of_week, subject, teacher FROM schedule WHERE lesson_number = {lesson_number} AND day_of_week = '{current_day_of_week}'")

        cursor.execute("""
            SELECT day_of_week, subject, teacher
            FROM schedule
            WHERE lesson_number = %s AND day_of_week = %s
        """, (lesson_number, current_day_of_week))
        schedule_info = cursor.fetchone()

        # Отладочный вывод в консоль
        print(f"Schedule Info: {schedule_info}")

        if schedule_info:
            updated_record = list(record) + list(schedule_info)
            updated_attendance_data.append(updated_record)
        else:
            updated_attendance_data.append(list(record) + [None, None, None])

    cursor.close()
    conn.close()

    # Отладочный вывод в консоль
    for row in updated_attendance_data:
        print(row)

    teacher_name = f"{current_user.last_name} {current_user.first_name} {current_user.middle_name}"
    return render_template('teacher_dashboard.html', attendance_data=updated_attendance_data, teacher_name=teacher_name, has_classes=has_classes)


@app.route('/teacher/add_student', methods=['GET', 'POST'])
@login_required
@role_required('teacher')
def add_student():
    form = AddStudentForm()
    form.class_id.choices = [(row[0], row[1]) for row in get_classes()]
    if form.validate_on_submit():
        print("Form validated successfully")
        last_name = form.last_name.data
        first_name = form.first_name.data
        middle_name = form.middle_name.data
        class_id = form.class_id.data
        photo = form.photo.data
        if photo and allowed_file(photo.filename):
            face_encoding = random.randint(100000, 999999)  # Генерация случайного числа для face_encoding
            filename = f"{face_encoding}.jpg"
            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            photo_path_for_profile = os.path.join(app.config['UPLOAD_FOLDER_FOR_PROFILES_S'], filename)
            # Сохранение файла в первую папку
            photo.save(photo_path)
            # Копирование файла во вторую папку
            shutil.copyfile(photo_path, photo_path_for_profile)
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO students (last_name, first_name, middle_name, face_encoding, class_id) VALUES (%s, %s, %s, %s, %s)",
                           (last_name, first_name, middle_name, face_encoding, class_id))
            conn.commit()
            cursor.close()
            conn.close()
            flash('Student added successfully', 'success')
            return redirect(url_for('my_class'))
        else:
            print("Invalid file format")
            flash('Invalid file format', 'danger')
    else:
        print("Form validation failed")
        for field, errors in form.errors.items():
            for error in errors:
                print(f"Error in {field}: {error}")
    return render_template('add_student.html', form=form)


def get_classes():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM classes WHERE name != 'No Class'")
    classes = cursor.fetchall()
    cursor.close()
    conn.close()
    return classes


@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username = %s", (form.username.data,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        if user and bcrypt.check_password_hash(user[2], form.password.data):
            login_user(User(user[0], user[1], user[2], user[3], user[4], user[5]))
            return redirect(url_for('index'))
        else:
            flash('Login Unsuccessful. Check username and password', 'danger')
    return render_template('login.html', form=form)


@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username = %s", (form.username.data,))
        user = cursor.fetchone()
        if user:
            flash('Username already exists', 'danger')
        else:
            role = form.role.data  # Получаем выбранную роль из формы
            cursor.execute("INSERT INTO users (username, password, role, last_name, first_name, middle_name) VALUES (%s, %s, %s, %s, %s, %s)",
                           (form.username.data, hashed_password, role, form.last_name.data, form.first_name.data, form.middle_name.data))
            conn.commit()
            flash('Account created successfully', 'success')
            return redirect(url_for('login'))
        cursor.close()
        conn.close()
    return render_template('register.html', form=form)




@app.route('/admin/teachers/edit/<int:teacher_id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_teacher(teacher_id):
    form = TeacherEditForm()
    form.class_id.choices = [(row[0], row[1]) for row in get_classes() if row[1] != 'No Class']

    if request.method == 'POST' and form.validate_on_submit():
        last_name = form.last_name.data
        first_name = form.first_name.data
        middle_name = form.middle_name.data
        email = form.email.data
        phone = form.phone.data
        education = form.education.data
        experience = form.experience.data
        has_classes = request.form.get('has_classes') == 'on'
        photo = form.photo.data

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            # Обновление учителя в таблице teachers
            if photo and allowed_file(photo.filename):
                face_encoding = random.randint(100000, 999999)  # Генерация случайного числа для face_encoding
                filename = f"{face_encoding}.jpg"
                photo_path = os.path.join(app.config['UPLOAD_FOLDER_FOR_PROFILES_T'], filename)
                photo.save(photo_path)
                cursor.execute("UPDATE teachers SET last_name = %s, first_name = %s, middle_name = %s, email = %s, phone = %s, education = %s, experience = %s, face_encoding = %s WHERE id = %s",
                               (last_name, first_name, middle_name, email, phone, education, experience, face_encoding, teacher_id))
            else:
                cursor.execute("UPDATE teachers SET last_name = %s, first_name = %s, middle_name = %s, email = %s, phone = %s, education = %s, experience = %s WHERE id = %s",
                               (last_name, first_name, middle_name, email, phone, education, experience, teacher_id))
            cursor.close()

            # Удаление старых связей с классами
            cursor = conn.cursor()
            cursor.execute("DELETE FROM teacher_classes WHERE teacher_id = %s", (teacher_id,))
            cursor.close()

            # Добавление новых связей с классами
            cursor = conn.cursor()
            if has_classes:
                class_ids = form.class_id.data
                for class_id in class_ids:
                    cursor.execute("INSERT INTO teacher_classes (teacher_id, class_id, has_classes) VALUES (%s, %s, %s)",
                                   (teacher_id, class_id, 1))
            else:
                cursor.execute("INSERT INTO teacher_classes (teacher_id, class_id, has_classes) VALUES (%s, %s, %s)",
                               (teacher_id, 0, 0))  # Используем 0 для обозначения отсутствия класса
            cursor.close()

            conn.commit()
            flash('Учитель обновлен успешно', 'success')
        except Exception as e:
            print(f"Error updating teacher: {str(e)}")
            flash(f'Ошибка при обновлении учителя: {str(e)}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('admin_teachers'))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT teachers.*, GROUP_CONCAT(classes.id ORDER BY classes.id SEPARATOR ', ') as class_ids, MAX(teacher_classes.has_classes) as has_classes
        FROM teachers
        LEFT JOIN teacher_classes ON teachers.id = teacher_classes.teacher_id
        LEFT JOIN classes ON teacher_classes.class_id = classes.id
        WHERE teachers.id = %s
        GROUP BY teachers.id
    """, (teacher_id,))
    teacher = cursor.fetchone()
    cursor.close()
    conn.close()

    if teacher:
        form.last_name.data = teacher[1]
        form.first_name.data = teacher[2]
        form.middle_name.data = teacher[3]
        form.email.data = teacher[4]
        form.phone.data = teacher[5]
        form.education.data = teacher[6]
        form.experience.data = teacher[7]
        class_ids = teacher[9].split(',') if teacher[9] else []
        form.class_id.data = [int(class_id) for class_id in class_ids if class_id.isdigit()]
        has_classes = teacher[10] == 1

        # Отладочная информация
        print(f"Teacher ID: {teacher_id}")
        print(f"Class IDs: {class_ids}")
        print(f"Has Classes: {has_classes}")
        print(f"Form Class ID Data: {form.class_id.data}")
        print(f"Teacher Data: {teacher}")

    return render_template('edit_teacher.html', form=form, teacher_id=teacher_id, teacher=teacher)


@app.route('/admin/teachers/delete/<int:teacher_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_teacher(teacher_id):
    print(f"Deleting teacher with ID: {teacher_id}")
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Начинаем транзакцию
        conn.start_transaction()

        # Удаление связанных записей из таблицы teacher_classes
        cursor.execute("DELETE FROM teacher_classes WHERE teacher_id = %s", (teacher_id,))
        print(f"Deleted related records from teacher_classes table with ID: {teacher_id}")

        # Удаление учителя из таблицы teachers
        cursor.execute("DELETE FROM teachers WHERE id = %s", (teacher_id,))
        print(f"Deleted teacher from teachers table with ID: {teacher_id}")

        # Удаление связанной записи из таблицы users
        cursor.execute("DELETE FROM users WHERE teacher_id = %s", (teacher_id,))
        print(f"Deleted teacher from users table with ID: {teacher_id}")

        # Подтверждаем транзакцию
        conn.commit()
        flash('Учитель удален успешно', 'success')

    except Exception as e:
        # Откатываем транзакцию в случае ошибки
        conn.rollback()
        flash(f'Ошибка при удалении учителя: {str(e)}', 'danger')
        print(f"Error deleting teacher: {str(e)}")

    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin_teachers'))


@app.route('/admin/students/delete/<int:student_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_student(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Начинаем транзакцию
        conn.start_transaction()

        # Удаление студента из таблицы students
        cursor.execute("DELETE FROM students WHERE id = %s", (student_id,))

        # Подтверждаем транзакцию
        conn.commit()
        flash('Ученик удален успешно', 'success')

    except Exception as e:
        # Откатываем транзакцию в случае ошибки
        conn.rollback()
        flash(f'Ошибка при удалении ученика: {str(e)}', 'danger')

    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin_students'))




@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/capture')
@login_required
@role_required('teacher')
def capture():
    global capture_active
    capture_active = True
    teacher_id = current_user.teacher_id
    threading.Thread(target=capture_and_recognize, args=(teacher_id,)).start()
    return jsonify({"status": "Capturing and recognizing faces..."})

@app.route('/stop_capture')
@login_required
@role_required('teacher')
def stop_capture():
    global capture_active
    capture_active = False
    return jsonify({"status": "Stopped capturing"})

@app.route('/get_attendance_data')
@login_required
@role_required('teacher')
def get_attendance_data():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name),
               classes.name, attendance.attendance_time, schedule.day_of_week,
               schedule.subject, schedule.teacher
        FROM attendance
        JOIN students ON attendance.student_id = students.id
        JOIN classes ON students.class_id = classes.id
        JOIN schedule ON attendance.lesson_number = schedule.lesson_number
                       AND DAYOFWEEK(attendance.attendance_time) =
                           CASE schedule.day_of_week
                               WHEN 'Понедельник' THEN 2
                               WHEN 'Вторник' THEN 3
                               WHEN 'Среда' THEN 4
                               WHEN 'Четверг' THEN 5
                               WHEN 'Пятница' THEN 6
                               WHEN 'Суббота' THEN 7
                               WHEN 'Воскресенье' THEN 1
                           END
        WHERE attendance.responsible_teacher_id = %s
    """, (current_user.teacher_id,))
    attendance_data = cursor.fetchall()
    cursor.close()
    conn.close()

    # Преобразование данных в список словарей для удобства
    attendance_list = []
    for name, class_name, attendance_time, day_of_week, subject, teacher in attendance_data:
        attendance_list.append({
            'name': name,
            'class': class_name,
            'attendance_time': attendance_time.strftime('%Y-%m-%d %H:%M:%S'),
            'day_of_week': day_of_week,
            'subject': subject,
            'teacher': teacher
        })

    return jsonify(attendance_list)


@app.route('/get_attendance_statistics')
@login_required
@role_required('admin')
def get_attendance_statistics():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DATE(attendance_time) as date, COUNT(*) as count FROM attendance GROUP BY DATE(attendance_time) ORDER BY date DESC LIMIT 7")
    statistics = cursor.fetchall()
    cursor.close()
    conn.close()

    # Преобразование данных в список словарей для удобства
    statistics_list = []
    for date, count in statistics:
        statistics_list.append({
            'date': date.strftime('%Y-%m-%d'),
            'count': count
        })

    return jsonify(statistics_list)

@app.route('/get_notifications')
@login_required
@role_required('admin')
def get_notifications():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT CONCAT('Новый ученик добавлен: ', students.last_name, ' ', students.first_name, ' ', students.middle_name) AS message
        FROM students
        ORDER BY students.id DESC
        LIMIT 10
    """)
    notifications = cursor.fetchall()
    cursor.close()
    conn.close()

    return jsonify([notification[0] for notification in notifications])


@app.route('/get_recent_activities')
@login_required
@role_required('admin')
def get_recent_activities():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name) AS student_name,
               classes.name AS class_name,
               teachers.last_name AS teacher_last_name,
               teachers.first_name AS teacher_first_name,
               teachers.middle_name AS teacher_middle_name,
               attendance.attendance_time
        FROM attendance
        JOIN students ON attendance.student_id = students.id
        JOIN classes ON students.class_id = classes.id
        JOIN teachers ON attendance.responsible_teacher_id = teachers.id
        ORDER BY attendance.attendance_time DESC
        LIMIT 10
    """)
    activities = cursor.fetchall()
    cursor.close()
    conn.close()

    # Преобразование данных в список словарей для удобства
    activities_list = []
    for student_name, class_name, teacher_last_name, teacher_first_name, teacher_middle_name, attendance_time in activities:
        activities_list.append({
            'student_name': student_name,
            'class_name': class_name,
            'teacher_name': f"Ответственный - {teacher_last_name} {teacher_first_name} {teacher_middle_name}",
            'attendance_time': attendance_time.strftime('%Y-%m-%d %H:%M:%S')  # Преобразование datetime в строку
        })

    return jsonify(activities_list)

@app.route('/search', methods=['POST'])
@login_required
@role_required('admin')
def search():
    query = request.form.get('query')
    results = []

    conn = get_db_connection()
    cursor = conn.cursor()

    # Поиск учеников
    cursor.execute("""
        SELECT CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name) AS student_name,
               classes.name AS class_name,
               teachers.last_name AS teacher_last_name,
               teachers.first_name AS teacher_first_name,
               teachers.middle_name AS teacher_middle_name,
               attendance.attendance_time
        FROM attendance
        JOIN students ON attendance.student_id = students.id
        JOIN classes ON students.class_id = classes.id
        JOIN teachers ON attendance.responsible_teacher_id = teachers.id
        WHERE CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name) LIKE %s
        OR classes.name LIKE %s
        OR CONCAT(teachers.last_name, ' ', teachers.first_name, ' ', teachers.middle_name) LIKE %s
    """, ('%' + query + '%', '%' + query + '%', '%' + query + '%'))
    search_results = cursor.fetchall()
    for student_name, class_name, teacher_last_name, teacher_first_name, teacher_middle_name, attendance_time in search_results:
        results.append({
            'student_name': student_name,
            'class_name': class_name,
            'teacher_name': f"{teacher_last_name} {teacher_first_name} {teacher_middle_name}",
            'attendance_time': attendance_time.strftime('%Y-%m-%d %H:%M:%S')  # Преобразование datetime в строку
        })

    cursor.close()
    conn.close()

    return jsonify(results)


@app.route('/get_settings')
@login_required
@role_required('admin')
def get_settings():
    # Логика для получения настроек
    settings = {
        "face_recognition_threshold": 0.6,
        "notification_frequency": "daily"
    }
    return jsonify(settings)

@app.route('/update_settings', methods=['POST'])
@login_required
@role_required('admin')
def update_settings():
    data = request.json
    # Логика для обновления настроек
    # Например, обновление порогового значения распознавания лиц
    face_recognition_threshold = data.get('face_recognition_threshold')
    notification_frequency = data.get('notification_frequency')
    # Здесь должна быть логика для сохранения настроек в базе данных
    return jsonify({"status": "success"})

@app.route('/reports')
@login_required
@role_required(['admin', 'teacher'])
def reports():
    conn = get_db_connection()
    cursor = conn.cursor()
    user_role = current_user.role

    if user_role == 'admin':
        # Получаем статистику посещаемости по всем классам для администратора
        cursor.execute("""
            SELECT classes.name, COUNT(attendance.id) as attendance_count
            FROM attendance
            JOIN students ON attendance.student_id = students.id
            JOIN classes ON students.class_id = classes.id
            GROUP BY classes.name
        """)
    elif user_role == 'teacher':
        # Получаем статистику посещаемости для классов, за которые отвечает учитель
        cursor.execute("""
            SELECT classes.name, COUNT(attendance.id) as attendance_count
            FROM attendance
            JOIN students ON attendance.student_id = students.id
            JOIN classes ON students.class_id = classes.id
            JOIN teacher_classes ON classes.id = teacher_classes.class_id
            WHERE teacher_classes.teacher_id = %s
            GROUP BY classes.name
        """, (current_user.teacher_id,))

    attendance_stats = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('reports.html', attendance_stats=attendance_stats)


import os
import pandas as pd
from flask import send_file
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

@app.route('/download_report')
@login_required
@role_required(['admin', 'teacher'])
def download_report():
    conn = get_db_connection()
    cursor = conn.cursor()
    user_role = current_user.role

    if user_role == 'admin':
        # Получаем данные о посещаемости для всех классов для администратора
        cursor.execute("""
            SELECT CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name) AS full_name,
                   students.face_encoding AS photo,
                   DATE(attendance.attendance_time) AS attendance_date,
                   classes.name AS class_name,
                   GROUP_CONCAT(DISTINCT schedule.subject ORDER BY schedule.subject SEPARATOR ', ') AS subjects
            FROM attendance
            JOIN students ON attendance.student_id = students.id
            JOIN classes ON students.class_id = classes.id
            JOIN schedule ON attendance.lesson_number = schedule.lesson_number
                           AND DAYOFWEEK(attendance.attendance_time) =
                               CASE schedule.day_of_week
                                   WHEN 'Понедельник' THEN 2
                                   WHEN 'Вторник' THEN 3
                                   WHEN 'Среда' THEN 4
                                   WHEN 'Четверг' THEN 5
                                   WHEN 'Пятница' THEN 6
                                   WHEN 'Суббота' THEN 7
                                   WHEN 'Воскресенье' THEN 1
                               END
            GROUP BY students.id, DATE(attendance.attendance_time)
        """)
    elif user_role == 'teacher':
        # Получаем данные о посещаемости для классов, за которые отвечает учитель
        cursor.execute("""
            SELECT CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name) AS full_name,
                   students.face_encoding AS photo,
                   DATE(attendance.attendance_time) AS attendance_date,
                   classes.name AS class_name,
                   GROUP_CONCAT(DISTINCT schedule.subject ORDER BY schedule.subject SEPARATOR ', ') AS subjects
            FROM attendance
            JOIN students ON attendance.student_id = students.id
            JOIN classes ON students.class_id = classes.id
            JOIN schedule ON attendance.lesson_number = schedule.lesson_number
                           AND DAYOFWEEK(attendance.attendance_time) =
                               CASE schedule.day_of_week
                                   WHEN 'Понедельник' THEN 2
                                   WHEN 'Вторник' THEN 3
                                   WHEN 'Среда' THEN 4
                                   WHEN 'Четверг' THEN 5
                                   WHEN 'Пятница' THEN 6
                                   WHEN 'Суббота' THEN 7
                                   WHEN 'Воскресенье' THEN 1
                               END
            JOIN teacher_classes ON classes.id = teacher_classes.class_id
            WHERE teacher_classes.teacher_id = %s
            GROUP BY students.id, DATE(attendance.attendance_time)
        """, (current_user.teacher_id,))

    attendance_data = cursor.fetchall()
    cursor.close()
    conn.close()

    # Создаем директорию для отчетов, если она не существует
    reports_dir = 'reports'
    os.makedirs(reports_dir, exist_ok=True)

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Заголовки для Excel
    headers = ['ФИО', 'Фото', 'Дата посещения', 'Класс', 'Предметы']
    ws.append(headers)

    # Добавляем данные в Excel
    for row in attendance_data:
        ws.append(list(row))

    # Вставка изображений
    for row in range(2, ws.max_row + 1):
        photo_filename = ws.cell(row=row, column=2).value
        if photo_filename:
            photo_path = os.path.join('static/students_photo', f"{photo_filename}.jpg")
            if os.path.exists(photo_path):
                img = ExcelImage(photo_path)
                # Устанавливаем размер изображения
                img.width, img.height = 80, 100
                # Вставляем изображение в ячейку B
                ws.add_image(img, f'B{row}')

    # Сохраняем Excel-файл
    report_path = os.path.join(reports_dir, 'attendance_report.xlsx')
    wb.save(report_path)

    return send_file(report_path, as_attachment=True)


@app.route('/my_class')
@login_required
@role_required('teacher')
def my_class():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Получаем классы учителя
    cursor.execute("SELECT class_id FROM teacher_classes WHERE teacher_id = %s", (current_user.teacher_id,))
    class_ids = cursor.fetchall()
    classes = []

    for class_id in class_ids:
        class_id = class_id[0]
        # Получаем название класса
        cursor.execute("SELECT name FROM classes WHERE id = %s", (class_id,))
        class_name = cursor.fetchone()[0]

        # Получаем учеников, привязанных к этому классу
        cursor.execute("""
            SELECT id, last_name, first_name, middle_name, face_encoding
            FROM students
            WHERE class_id = %s
        """, (class_id,))
        students = cursor.fetchall()

        classes.append({'class_id': class_id, 'class_name': class_name, 'students': students})

    cursor.close()
    conn.close()

    return render_template('my_class.html', classes=classes)


@app.route('/teacher_profile')
@login_required
@role_required('teacher')
def teacher_profile():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT t.last_name, t.first_name, t.middle_name, GROUP_CONCAT(c.name ORDER BY c.name SEPARATOR ', ') AS class_names, t.email, t.phone, t.education, t.experience, t.face_encoding
        FROM teachers t
        JOIN teacher_classes tc ON t.id = tc.teacher_id
        JOIN classes c ON tc.class_id = c.id
        WHERE t.id = %s
        GROUP BY t.id
    """, (current_user.teacher_id,))
    teacher_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if teacher_data:
        teacher_name = f"{teacher_data[0]} {teacher_data[1]} {teacher_data[2]}"
        teacher_class = teacher_data[3]
        teacher_email = teacher_data[4]
        teacher_phone = teacher_data[5]
        teacher_education = teacher_data[6]
        teacher_experience = teacher_data[7]
        teacher_photo = teacher_data[8]
    else:
        teacher_name = teacher_class = teacher_email = teacher_phone = teacher_education = teacher_experience = teacher_photo = ""

    print("Fetced data: ", teacher_data)

    return render_template('teacher_profile.html',
                           teacher_name=teacher_name,
                           teacher_class=teacher_class,
                           teacher_email=teacher_email,
                           teacher_phone=teacher_phone,
                           teacher_education=teacher_education,
                           teacher_experience=teacher_experience,
                           teacher_photo = teacher_photo)

@app.route('/class/<int:class_id>')
@login_required
@role_required('admin')
def class_details(class_id):
    conn = get_db_connection()

    try:
        # Получаем информацию о классе
        cursor = conn.cursor(buffered=True)
        cursor.execute("SELECT name FROM classes WHERE id = %s", (class_id,))
        class_name = cursor.fetchone()[0]
        cursor.close()

        # Получаем классного руководителя
        teachers = get_teachers()
        teacher = next((t for t in teachers if t[10] == class_id and t[8] == 1), None)

        # Получаем учеников класса
        students = get_students()
        class_students = [s for s in students if s[5] == class_id]

    finally:
        conn.close()

    return render_template('class_details.html', class_name=class_name, teacher=teacher, students=class_students)


def random_color():
    return f"{random.randint(0, 255)}, {random.randint(0, 255)}, {random.randint(0, 255)}"

@app.context_processor
def inject_random_color():
    return dict(random_color=random_color)

@app.route('/admin/students/edit/<int:student_id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_student(student_id):
    form = AddStudentForm()
    form.class_id.choices = [(row[0], row[1]) for row in get_classes()]

    if request.method == 'POST' and form.validate_on_submit():
        last_name = form.last_name.data
        first_name = form.first_name.data
        middle_name = form.middle_name.data
        class_id = form.class_id.data
        photo = form.photo.data

        if photo and allowed_file(photo.filename):
            face_encoding = random.randint(100000, 999999)  # Генерация случайного числа для face_encoding
            filename = f"{face_encoding}.jpg"
            photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            photo_path_for_profile = os.path.join(app.config['UPLOAD_FOLDER_FOR_PROFILE'], filename)
            # Сохранение файла в первую папку
            photo.save(photo_path)
            # Копирование файла во вторую папку
            shutil.copyfile(photo_path, photo_path_for_profile)
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE students SET last_name = %s, first_name = %s, middle_name = %s, face_encoding = %s, class_id = %s WHERE id = %s",
                           (last_name, first_name, middle_name, face_encoding, class_id, student_id))
            conn.commit()
            cursor.close()
            conn.close()
            flash('Ученик обновлен успешно', 'success')
            return redirect(url_for('admin_students'))
        else:
            flash('Неверный формат файла', 'danger')

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE id = %s", (student_id,))
    student = cursor.fetchone()
    cursor.close()
    conn.close()

    if student:
        form.last_name.data = student[1]
        form.first_name.data = student[2]
        form.middle_name.data = student[3]
        form.class_id.data = student[5]

    return render_template('edit_student.html', form=form, student_id=student_id)

def get_teachers_with_combined_classes():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT teachers.id, teachers.last_name, teachers.first_name, teachers.middle_name,
               users.username, users.password, users.role,
               GROUP_CONCAT(classes.name ORDER BY classes.name SEPARATOR ', ') as class_names,
               MAX(teacher_classes.has_classes) as has_classes,
               teachers.face_encoding,
               teachers.email, teachers.phone, teachers.education, teachers.experience
        FROM teachers
        JOIN users ON teachers.id = users.teacher_id
        LEFT JOIN teacher_classes ON teachers.id = teacher_classes.teacher_id
        LEFT JOIN classes ON teacher_classes.class_id = classes.id
        GROUP BY teachers.id, users.username, users.password, users.role
    """)
    teachers = cursor.fetchall()
    cursor.close()
    conn.close()

    # Отладочная информация
    print("Teachers data fetched from the database:")
    for teacher in teachers:
        print(teacher)

    return teachers

@app.route('/schedule')
@login_required
@role_required('admin')
def schedule():
    return render_template('schedule.html')

@app.route('/upload_schedule', methods=['POST'])
@login_required
@role_required('admin')
def upload_schedule():
    if 'file' not in request.files:
        app.logger.error("No file part in the request")
        return jsonify({"status": "error", "message": "No file part"})

    file = request.files['file']
    if file.filename == '':
        app.logger.error("No selected file")
        return jsonify({"status": "error", "message": "No selected file"})

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        app.logger.info(f"File saved to {file_path}")

        # Parse the schedule from the uploaded file
        schedule_data = parse_schedule(file_path)
        if schedule_data:
            app.logger.info("Schedule parsed successfully")
            save_schedule_to_db(schedule_data)
            return jsonify({"status": "success", "schedule": schedule_data})
        else:
            app.logger.error("Failed to parse schedule")
            return jsonify({"status": "error", "message": "Failed to parse schedule"})
    else:
        app.logger.error("Invalid file type")
        return jsonify({"status": "error", "message": "Invalid file type"})

def save_schedule_to_db(schedule_data):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Удаление старых данных расписания
    cursor.execute("DELETE FROM schedule")
    conn.commit()

    # Вставка новых данных расписания
    for lesson in schedule_data:
        cursor.execute("""
            INSERT INTO schedule (date, day_of_week, lesson_number, time, class_name, subject, teacher, room)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (lesson['date'], lesson['day_of_week'], lesson['lesson_number'], lesson['time'], lesson['class_name'], lesson['subject'], lesson['teacher'], lesson['room']))

    conn.commit()
    cursor.close()
    conn.close()
    print("Schedule saved to database successfully")



def parse_schedule(file_path):
    class Schedule:
        def __init__(self):
            self.data = defaultdict(list)  # хранит расписание: {класс: список занятий}
            self.current_date = None
            self.current_day_of_week = None

        def add_lesson(self, class_name, lesson_number, lesson_time, subject, teacher, room):
            self.data[class_name].append({
                "date": self.current_date,
                "day_of_week": self.current_day_of_week,
                "lesson_number": lesson_number,
                "time": lesson_time,
                "subject": subject,
                "teacher": teacher,
                "room": room
            })

        def to_dataframe(self):
            # Создаем список всех занятий
            all_lessons = []
            for class_name, lessons in self.data.items():
                for lesson in lessons:
                    lesson['class_name'] = class_name
                    all_lessons.append(lesson)

            # Создаем DataFrame из списка занятий
            df = pd.DataFrame(all_lessons)
            return df

        def __repr__(self):
            return str(dict(self.data))

    try:
        # Чтение файла
        df = pd.read_excel(file_path, sheet_name='Расписание')
        print("Excel file read successfully")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

    schedule = Schedule()

    # Парсинг данных
    for index, row in df.iterrows():
        date_time = row['Дата / Время']

        # Проверка на наличие даты и дня недели
        date_match = re.match(r'(\d{2}\.\d{2}\.\d{4}) \((.+)\)', date_time)
        if date_match:
            schedule.current_date = date_match.group(1)
            schedule.current_day_of_week = date_match.group(2)
            print(f"Date and day of week set: {schedule.current_date}, {schedule.current_day_of_week}")
            continue  # Пропуск строки с датой и днем недели

        # Разделение строки на номер урока и время
        lesson_info = date_time.split('\n')
        lesson_number = int(lesson_info[0].split()[0])  # Извлекаем только цифру
        lesson_time = lesson_info[1]

        for class_name in row.index[2:]:
            lesson_info = row[class_name]
            if not pd.isna(lesson_info):
                lesson_parts = lesson_info.split(", ")
                subject = lesson_parts[0]
                teacher = lesson_parts[-1]
                room = lesson_parts[-2].replace("Каб.", "").strip()
                schedule.add_lesson(class_name, lesson_number, lesson_time, subject, teacher, room)
                print(f"Lesson added: {class_name}, {lesson_number}, {lesson_time}, {subject}, {teacher}, {room}")

    # Преобразование расписания в DataFrame
    schedule_df = schedule.to_dataframe()

    # Настройка отображения всех строк и столбцов
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)  # Устанавливаем ширину экрана на максимальную

    # Вывод расписания в виде таблицы
    print("Schedule parsed successfully")
    return schedule_df.to_dict(orient='records')



@app.route('/get_schedule', methods=['GET'])
@login_required
@role_required('admin')
def get_schedule():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM schedule")
    schedule_data = cursor.fetchall()
    cursor.close()
    conn.close()

    if schedule_data:
        return jsonify({"status": "success", "schedule": schedule_data})
    else:
        return jsonify({"status": "error", "message": "No schedule data found"})


@app.route('/get_filters', methods=['GET'])
@login_required
@role_required('admin')
def get_filters():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Получение уникальных дней недели
    cursor.execute("SELECT DISTINCT day_of_week FROM schedule")
    days_of_week = [row['day_of_week'] for row in cursor.fetchall()]

    # Получение уникальных классов
    cursor.execute("SELECT DISTINCT class_name FROM schedule")
    classes = [row['class_name'] for row in cursor.fetchall()]

    # Получение уникальных времен
    cursor.execute("SELECT DISTINCT time FROM schedule")
    times = [row['time'] for row in cursor.fetchall()]

    # Получение уникальных предметов
    cursor.execute("SELECT DISTINCT subject FROM schedule")
    subjects = [row['subject'] for row in cursor.fetchall()]

    # Получение уникальных учителей
    cursor.execute("SELECT DISTINCT teacher FROM schedule")
    teachers = [row['teacher'] for row in cursor.fetchall()]

    cursor.close()
    conn.close()

    return jsonify({
        "status": "success",
        "filters": {
            "daysOfWeek": days_of_week,
            "classes": classes,
            "times": times,
            "subjects": subjects,
            "teachers": teachers,
        }
    })



@app.route('/delete_schedule', methods=['POST'])
@login_required
@role_required('admin')
def delete_schedule():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM schedule")
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"status": "success", "message": "Расписание успешно удалено"})

@app.route('/teacher_schedule')
@login_required
@role_required('teacher')
def teacher_schedule():
    return render_template('teacher_schedule.html')


@app.route('/get_teacher_schedule', methods=['GET'])
@login_required
@role_required('teacher')
def get_teacher_schedule():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    teacher_id = current_user.teacher_id

    # Получаем классы, к которым привязан учитель
    cursor.execute("""
        SELECT classes.name AS class_name
        FROM teacher_classes
        JOIN classes ON teacher_classes.class_id = classes.id
        WHERE teacher_classes.teacher_id = %s AND teacher_classes.has_classes = 1
    """, (teacher_id,))
    class_names = [row['class_name'] for row in cursor.fetchall()]

    if not class_names:
        return jsonify({"status": "error", "message": "У вас нет привязанных классов."})

    # Получаем расписание для этих классов
    placeholders = ', '.join(['%s'] * len(class_names))
    cursor.execute(f"""
        SELECT *
        FROM schedule
        WHERE class_name IN ({placeholders})
    """, tuple(class_names))
    schedule_data = cursor.fetchall()
    cursor.close()
    conn.close()

    if schedule_data:
        return jsonify({"status": "success", "schedule": schedule_data})
    else:
        return jsonify({"status": "error", "message": "Расписание не найдено."})

@app.route('/get_teacher_schedule_filters', methods=['GET'])
@login_required
@role_required('teacher')
def get_teacher_schedule_filters():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    teacher_id = current_user.teacher_id

    # Получаем классы, к которым привязан учитель
    cursor.execute("""
        SELECT classes.name AS class_name
        FROM teacher_classes
        JOIN classes ON teacher_classes.class_id = classes.id
        WHERE teacher_classes.teacher_id = %s AND teacher_classes.has_classes = 1
    """, (teacher_id,))
    class_names = [row['class_name'] for row in cursor.fetchall()]

    if not class_names:
        return jsonify({"status": "error", "message": "У вас нет привязанных классов."})

    # Получаем уникальные дни недели, временные интервалы, предметы и учителей для этих классов
    cursor.execute("""
        SELECT DISTINCT day_of_week
        FROM schedule
        WHERE class_name IN (%s)
    """ % ','.join(['%s'] * len(class_names)), tuple(class_names))
    days_of_week = [row['day_of_week'] for row in cursor.fetchall()]

    cursor.execute("""
        SELECT DISTINCT time
        FROM schedule
        WHERE class_name IN (%s)
    """ % ','.join(['%s'] * len(class_names)), tuple(class_names))
    times = [row['time'] for row in cursor.fetchall()]

    cursor.execute("""
        SELECT DISTINCT subject
        FROM schedule
        WHERE class_name IN (%s)
    """ % ','.join(['%s'] * len(class_names)), tuple(class_names))
    subjects = [row['subject'] for row in cursor.fetchall()]

    cursor.execute("""
        SELECT DISTINCT teacher
        FROM schedule
        WHERE class_name IN (%s)
    """ % ','.join(['%s'] * len(class_names)), tuple(class_names))
    teachers = [row['teacher'] for row in cursor.fetchall()]

    cursor.close()
    conn.close()

    return jsonify({
        "status": "success",
        "filters": {
            "daysOfWeek": days_of_week,
            "classes": class_names,
            "times": times,
            "subjects": subjects,
            "teachers": teachers,
        }
    })

@app.route('/attendance_grid')
@login_required
@role_required('teacher')
def attendance_grid():
    from datetime import datetime

    # Получаем текущую дату и день недели
    current_date = datetime.now().strftime('%Y-%m-%d')
    current_day_of_week = datetime.now().strftime('%A')

    # Преобразуем текущий день недели на русский язык
    days_mapping = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье'
    }
    current_day_of_week_ru = days_mapping.get(current_day_of_week, current_day_of_week)

    conn = get_db_connection()
    cursor = conn.cursor()

    # Получаем учеников, за которых отвечал учитель в текущий день
    cursor.execute("""
        SELECT DISTINCT students.id, CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name) AS student_name, classes.name AS class_name
        FROM attendance
        JOIN students ON attendance.student_id = students.id
        JOIN classes ON students.class_id = classes.id
        WHERE DATE(attendance.attendance_time) = %s AND attendance.responsible_teacher_id = %s
    """, (current_date, current_user.teacher_id))
    students = cursor.fetchall()

    # Отладочная информация: выводим данные об учениках
    print("Students data fetched from the database:")
    for student in students:
        print(student)

    # Получаем предметы, которые ведет учитель в текущий день для каждого класса
    subjects_by_class = {}
    for student in students:
        class_name = student[2]
        cursor.execute("""
            SELECT DISTINCT lesson_number, subject
            FROM schedule
            WHERE day_of_week = %s AND class_name = %s
        """, (current_day_of_week_ru, class_name))
        subjects = cursor.fetchall()
        subjects_by_class[class_name] = {subject[0]: subject[1] for subject in subjects}

    # Отладочная информация: выводим данные о предметах
    print("Subjects data fetched from the database:")
    for class_name, subjects in subjects_by_class.items():
        print(f"Class Name {class_name}: {subjects}")

    # Получаем данные о посещаемости
    attendance_data = {}
    for student in students:
        student_id = student[0]
        class_name = student[2]
        attendance_data[student_id] = {lesson_number: False for lesson_number in subjects_by_class[class_name].keys()}

    cursor.execute("""
        SELECT student_id, lesson_number
        FROM attendance
        WHERE DATE(attendance.attendance_time) = %s AND attendance.responsible_teacher_id = %s
    """, (current_date, current_user.teacher_id))
    attendance_records = cursor.fetchall()

    # Отладочная информация: выводим данные о посещаемости
    print("Attendance records fetched from the database:")
    for record in attendance_records:
        print(record)

    for record in attendance_records:
        student_id, lesson_number = record
        if student_id in attendance_data and lesson_number in attendance_data[student_id]:
            attendance_data[student_id][lesson_number] = True
        else:
            print(f"Warning: Student ID {student_id} or lesson number {lesson_number} not found in attendance_data.")

    cursor.close()
    conn.close()

    # Отладочная информация
    print("Данные учеников перед рендерингом:", students)
    print("Данные предметов по классам перед рендерингом:", subjects_by_class)

    # Проверка, существует ли класс первого ученика в subjects_by_class
    if students:
        first_student_class = students[0][2]
        print(f"Класс первого ученика: {first_student_class}")
        if first_student_class in subjects_by_class:
            print(f"Предметы для {first_student_class}: {subjects_by_class[first_student_class]}")
        else:
            print(f"Предупреждение: {first_student_class} не найден в subjects_by_class")
    else:
        print("Предупреждение: Нет данных об учениках")

    return render_template('attendance_grid.html', students=students, subjects_by_class=subjects_by_class, attendance_data=attendance_data, current_date=current_date)

@app.route('/get_attendance_stats/<int:student_id>')
@login_required
@role_required('teacher')
def get_attendance_stats(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DATE(attendance_time) AS date, COUNT(*) AS count
        FROM attendance
        WHERE student_id = %s
        GROUP BY DATE(attendance_time)
        ORDER BY date DESC
        LIMIT 7
    """, (student_id,))
    stats = cursor.fetchall()
    cursor.close()
    conn.close()

    labels = [row[0].strftime('%Y-%m-%d') for row in stats]
    data = [row[1] for row in stats]

    return jsonify({'labels': labels, 'data': data})

from datetime import datetime

@app.route('/check_today_attendance', methods=['GET'])
@login_required
@role_required('teacher')
def check_today_attendance():
    conn = get_db_connection()
    cursor = conn.cursor()
    current_date = datetime.now().strftime('%Y-%m-%d')
    cursor.execute("""
        SELECT COUNT(*)
        FROM attendance
        WHERE DATE(attendance_time) = %s AND responsible_teacher_id = %s
    """, (current_date, current_user.teacher_id))
    count = cursor.fetchone()[0]
    cursor.close()
    conn.close()

    has_attendance = count > 0
    return jsonify({"has_attendance": has_attendance})



@app.route('/register_parent', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def register_parent():
    form = ParentRegistrationForm()
    form.class_id.choices = [(row[0], row[1]) for row in get_classes()]
    last_name = form.last_name.data
    first_name = form.first_name.data
    middle_name = form.middle_name.data
    def get_students_by_class(class_id):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, last_name, first_name, middle_name FROM students WHERE class_id = %s", (class_id,))
        students = cursor.fetchall()
        cursor.close()
        conn.close()
        return students

    if request.method == 'POST':
        class_id = request.form.get('class_id')
        print(f"Selected class_id: {class_id}")  # Отладочная печать

        if class_id is not None:
            # Получаем учеников для выбранного класса
            students = get_students_by_class(class_id)
            form.student_id.choices = [(row[0], f"{row[1]} {row[2]} {row[3]}") for row in students]
        else:
            flash('Пожалуйста, выберите класс', 'danger')
            return redirect(url_for('register_parent'))

        if form.validate_on_submit():
            hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
            conn = get_db_connection()
            cursor = conn.cursor()

            # Проверка на уникальность имени пользователя
            cursor.execute("SELECT * FROM users WHERE username = %s", (form.username.data,))
            existing_user = cursor.fetchone()
            if existing_user:
                flash('Имя пользователя уже существует', 'danger')
                return redirect(url_for('register_parent'))

            # Вставка нового родителя в таблицу parents
            cursor.execute("INSERT INTO parents (last_name, first_name, middle_name) VALUES (%s, %s, %s)",
                       (last_name, first_name, middle_name))
            parent_id = cursor.lastrowid
            # Вставка нового родителя в таблицу users
            cursor.execute("INSERT INTO users (username, password, role, parent_id) VALUES (%s, %s, %s, %s)",
                           (form.username.data, hashed_password, 'parent', parent_id))
           

            # Убедимся, что form.student_id.data - это список
            student_ids = form.student_id.data
            if not isinstance(student_ids, list):
                student_ids = [student_ids]

            # Вставка связей в таблицу parent_student_relations для каждого выбранного ученика
            for student_id in student_ids:
                cursor.execute("INSERT INTO parent_student_relations (parent_id, student_id) VALUES (%s, %s)",
                               (parent_id, student_id))

            conn.commit()
            cursor.close()
            conn.close()

            flash('Родитель успешно зарегистрирован', 'success')
            return redirect(url_for('admin_dashboard'))

    return render_template('register_parent.html', form=form)



@app.route('/get_students_by_class', methods=['POST'])
@login_required
@role_required('admin')
def get_students_by_class():
    # Получаем class_id из запроса
    class_id = request.form.get('class_id')
    print(f"Received class_id: {class_id}")  # Отладочная информация

    if class_id is None:
        print("Error: class_id is None")  # Отладочная информация
        return jsonify({"error": "Invalid class_id"}), 400

    try:
        # Устанавливаем соединение с базой данных
        conn = get_db_connection()
        cursor = conn.cursor()
        print("Database connection established")  # Отладочная информация

        # Выполняем SQL-запрос для получения учеников по class_id
        cursor.execute("SELECT id, last_name, first_name, middle_name FROM students WHERE class_id = %s", (class_id,))
        students = cursor.fetchall()
        print(f"Students fetched from database: {students}")  # Отладочная информация

        # Формируем список учеников для отправки в JSON формате
        students_list = [{'id': student[0], 'last_name': student[1], 'first_name': student[2], 'middle_name': student[3]} for student in students]
        print(f"Prepared students list for JSON response: {students_list}")  # Отладочная информация

        return jsonify(students_list)

    except Exception as e:
        print(f"Error fetching students: {e}")  # Отладочная информация
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()
        print("Database connection closed")  # Отладочная информация


@app.route('/parent_dashboard', methods=['GET', 'POST'])
@login_required
@role_required('parent')
def parent_dashboard():
    parent_id = current_user.parent_id

    # Получаем детей родителя
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT students.id, CONCAT(students.last_name, ' ', students.first_name, ' ', students.middle_name) AS full_name, classes.name AS class_name
        FROM parent_student_relations
        JOIN students ON parent_student_relations.student_id = students.id
        JOIN classes ON students.class_id = classes.id
        WHERE parent_student_relations.parent_id = %s
    """, (parent_id,))
    children = cursor.fetchall()

    # Отладочный вывод
    print("Данные о детях, полученные из базы данных:", children)

    # Получаем посещаемость для выбранного ребенка
    selected_child_id = request.args.get('child_id')
    attendance_data = {}
    subjects_by_class = {}
    current_date = datetime.now().strftime('%Y-%m-%d')
    current_day_of_week = datetime.now().strftime('%A')

    # Преобразуем текущий день недели на русский язык
    days_mapping = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье'
    }
    current_day_of_week_ru = days_mapping.get(current_day_of_week, current_day_of_week)

    # Отладочный вывод
    print(f"Текущая дата: {current_date}")
    print(f"Текущий день недели (на английском): {current_day_of_week}")
    print(f"Текущий день недели (на русском): {current_day_of_week_ru}")

    # Проверка, является ли сегодня выходным днем
    if current_day_of_week_ru in ['Суббота', 'Воскресенье']:
        no_schedule_message = f"Сегодня {current_day_of_week_ru}, занятий нет."
        return render_template('parent_dashboard.html', children=children, no_schedule_message=no_schedule_message, selected_child_id=selected_child_id, current_date=current_date)

    if selected_child_id:
        # Получаем данные о предметах из расписания для класса ученика на сегодня
        cursor.execute("""
            SELECT DISTINCT lesson_number, subject
            FROM schedule
            WHERE class_name = %s
            AND day_of_week = %s
        """, (children[int(selected_child_id)-1][2], current_day_of_week_ru))
        subjects = cursor.fetchall()

        # Отладочный вывод
        print(f"Данные о предметах для класса {children[int(selected_child_id)-1][2]} на {current_day_of_week_ru}: {subjects}")

        # Создаем словарь предметов по номеру урока
        subjects_by_class[selected_child_id] = {subject[0]: subject[1] for subject in subjects}

        # Отладочный вывод
        print("Словарь предметов по номеру урока:", subjects_by_class)

        # Инициализируем данные о посещаемости
        attendance_data[selected_child_id] = {lesson_number: False for lesson_number in subjects_by_class[selected_child_id].keys()}

        # Получаем данные о посещаемости
        cursor.execute("""
            SELECT lesson_number
            FROM attendance
            WHERE student_id = %s AND DATE(attendance_time) = %s
        """, (selected_child_id, current_date))
        attendance_records = cursor.fetchall()

        # Отладочный вывод
        print("Записи о посещаемости для ученика:", attendance_records)

        for record in attendance_records:
            lesson_number = record[0]
            if lesson_number in attendance_data[selected_child_id]:
                attendance_data[selected_child_id][lesson_number] = True

        # Отладочный вывод
        print("Данные о посещаемости:", attendance_data)

    cursor.close()
    conn.close()

    return render_template('parent_dashboard.html', children=children, attendance_data=attendance_data, subjects_by_class=subjects_by_class, selected_child_id=selected_child_id, current_date=current_date)

@app.route('/get_attendance_stats_for_student/<int:student_id>')
@login_required
@role_required('parent')
def get_attendance_stats_for_student(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DATE(attendance_time) AS date, COUNT(*) AS count
        FROM attendance
        WHERE student_id = %s
        GROUP BY DATE(attendance_time)
        ORDER BY date DESC
        LIMIT 7
    """, (student_id,))
    stats = cursor.fetchall()
    cursor.close()
    conn.close()

    labels = [row[0].strftime('%Y-%m-%d') for row in stats]
    data = [row[1] for row in stats]

    return jsonify({'labels': labels, 'data': data})


if __name__ == '__main__':
    socketio.run(app, debug=True, allow_unsafe_werkzeug=True)
